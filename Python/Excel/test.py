from openpyxl import workbook, load_workbook

wb = load_workbook('file.xlsx')

ws = wb.active

print(ws['A1'].value)
